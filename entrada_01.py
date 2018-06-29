import openpyxl
import datetime
import time

# import notify2

# notify2.init('foo')
# mensagem = notify2.Notification('Parabéns', 'Os dados foram incluídos com sucesso')





def __init__():
    
    planilha_origem = 'PlanilhadeEstoque2.xlsx'
    planilha_destino = 'marmileve.xlsx'

    ab = openpyxl.load_workbook(planilha_destino)

    print('As abas da planilha de destino são \n' + str(ab.sheetnames))
    listaclientes = []

    print('''
      __  __          _____  __  __ _____ _      ________      ________ 
     |  \/  |   /\   |  __ \|  \/  |_   _| |    |  ____\ \    / |  ____|
     | \  / |  /  \  | |__) | \  / | | | | |    | |__   \ \  / /| |__   
     | |\/| | / /\ \ |  _  /| |\/| | | | | |    |  __|   \ \/ / |  __|  
     | |  | |/ ____ \| | \ \| |  | |_| |_| |____| |____   \  /  | |____ 
     |_|  |_/_/    \_|_|  \_|_|  |_|_____|______|______|   \/   |______|


    ''')

def adicionarpedido():
    ab = openpyxl.load_workbook('marmileve.xlsx')
    sheet_pedidos = ab['Planilha1']
    sheet_tabela = ab['Tabela']
    sheet_estoque = ab['Estoque']
    sheet_clientes = ab['clientes']

    nropedido = sheet_pedidos['D':'D'][-1].value + 1
    print('O número do pedido é ' + str(nropedido))

    data = datetime.date.today()

    #Inicio Checagem/Cadastro de Clientes.

    cliente = input('Digite o nome do cliente: ')

    if checar_cliente(sheet_clientes, cliente) != 1:
        checar_cliente(sheet_clientes, cliente)
    else:
        pass
    
    # Fim da parte de Cliente.
    # Início da parte de pedidos.
    # Vale transformar isso numa função.

    print('\nAgora vamos preencher o pedido do cliente\n')

    item = 1
    resposta = 'sim'
    pedidos = []
    while resposta == 'sim':

        prato = int(input('Digite o número do prato: '))

        pratonome = {}
        for i in range(2, 22):
            pratonome[(sheet_tabela.cell(row=i, column=1).value)] = (sheet_tabela.cell(row=i, column=2).value)

        tamanho = input('Digite o tamanho do prato: ').lower()
        while tamanho not in ['p', 'g']:
            print('O tamanho não existe. Por favor digite de novo.')
            tamanho = input('Digite o tamanho do prato: ').lower()

        quantidade = int(input('Digite a quantidade de marmitas: '))

        for i in range(0, quantidade):
            pedidos.append([prato, tamanho, item])
            item += 1

        resposta = input('Você deseja adicionar mais algum prato(Sim ou nao)? ').lower()

    print(pedidos)
    ultimalinha = sheet_pedidos.max_row
    for i in pedidos:
        sheet_pedidos.cell(row=ultimalinha + i[2], column=1).value = data.month
        sheet_pedidos.cell(row=ultimalinha + i[2], column=2).value = data.year
        sheet_pedidos.cell(row=ultimalinha + i[2], column=3).value = data.isocalendar()[1]
        sheet_pedidos.cell(row=ultimalinha + i[2], column=4).value = nropedido
        sheet_pedidos.cell(row=ultimalinha + i[2], column=5).value = float((str(nropedido) + '.' + str(i[2])))
        sheet_pedidos.cell(row=ultimalinha + i[2], column=6).value = data
        sheet_pedidos.cell(row=ultimalinha + i[2], column=7).value = cliente
        sheet_pedidos.cell(row=ultimalinha + i[2], column=8).value = prato
        sheet_pedidos.cell(row=ultimalinha + i[2], column=9).value = pratonome[i[0]]
        sheet_pedidos.cell(row=ultimalinha + i[2], column=10).value = i[1]
        
        
        

    listapreco = {'p' : 17, 'g' : 19, 's' : 8}
    preco = listapreco[tamanho.lower()]


    # fim da parte de adicionar os pedidos.

    # Início da checagem de estoque

    checarestoque(sheet_estoque, pedidos, nropedido, data)



    # Fim da checagem de estoque

    fim = input('Deseja adicionar mais algum pedido?(S ou N) ').lower()
    
    if fim == 's':
        adicionarpedido()
    else:
        ab.save('marmileve.xlsx')
        print('\nPedido adicionado com sucesso')
        time.sleep(5)
    
def checarestoque(sheet_estoque, pedidos, nropedido, data):
    print('\nChecando disponibilidade no estoque')

    for a in range(0, len(pedidos)):
        t = 0
        i = 1
        while t < 1:
            if (sheet_estoque.cell(row=i, column=2).value == pedidos[a][0] and 
            sheet_estoque.cell(row=i, column=4).value.lower() == pedidos[a][1]):
                ultimadata = sheet_estoque.cell(row=i, column=6).value
                if sheet_estoque.cell(row=i, column=5).value == None:
                    sheet_estoque.cell(row=i, column=5).value = str(str(nropedido) + '.' + str(pedidos[a][2]))
                    print(str(str(nropedido) + '.' + str(pedidos[a][2])) + ' está disponível no dia ' + str(ultimadata))
                    
                    t += 1

            i += 1
            g = 0
            ultimalinha = sheet_estoque.max_row
            if sheet_estoque.cell(row=i, column=2).value == None:
                sheet_estoque.cell(row=i,column=1).value = ultimalinha
                sheet_estoque.cell(row=i,column=2).value = pedidos[a][0]
                sheet_estoque.cell(row=i,column=4).value = pedidos[a][1]
                for x in sheet_estoque['B':'F']:
                    if x[0].value == pedidos[a][0] and x[2].value == pedidos[a][1] and x[5] > datetime.date.today():
                        if ultimadata < x[5]:
                            g = 0
                            ultimadata == x[5]
                        else:
                            g += 1
                            
                sheet_estoque.cell(row=i,column=5).value = (str(nropedido) + '.' + str(pedidos[a][2]))
                sheet_estoque.cell(row=i,column=6).value = (ultimadata) + datetime.timedelta(days=1)
                print(str(str(nropedido) + '.' + str(pedidos[a][2])) + ' está disponível no dia ' + str(sheet_estoque.cell(row=i,column=6).value))
                break
            


    print('Pedidos conferidos')


def checar_cliente(sheet_clientes, cliente):

    ultimo = sheet_clientes.max_row+1
    
    listaclientes = []
    for i in range(2, sheet_clientes.max_row + 1):
        listaclientes.append(sheet_clientes.cell(i, column=2).value.lower())
    
    if cliente.lower() not in listaclientes:
        print('''
            \nEste cliente ainda no foi cadastrado na base de dados.
            \n
            \nEstamos lhe redirecionando para a tela de cadastro de clientes
            ''')

        userid = sheet_clientes.max_row

        time.sleep(2)
        
        print('''

        Por favor digite as informações referentes ao cliente:

            '''
              )
        
        endereco = input('\nDigite o Endereço do cliente: ')
        numero = input('\nDigite o numero: ')
        complemento = input('\nDigite o complemento: ')
        ptoreferencia = input('\nDigite o ponto de referencia: ')
        bairro = input('\nDigite o bairro: ')
        cidade = input('\nDigite a cidade: ')
        telefone = input('\nDigite o telefone: ')
        email = input('\nDigite o email: ')
        aniversario = input('\nDigite a data de aniversário: ')
        cpf = input('\nDigite o numero de CPF: ')

        sheet_clientes.cell(row=ultimo,column=1).value = userid
        sheet_clientes.cell(row=ultimo,column=2).value = cliente
        sheet_clientes.cell(row=ultimo,column=3).value = endereco
        sheet_clientes.cell(row=ultimo,column=4).value = numero
        sheet_clientes.cell(row=ultimo,column=5).value = complemento
        sheet_clientes.cell(row=ultimo,column=6).value = ptoreferencia
        sheet_clientes.cell(row=ultimo,column=7).value = bairro
        sheet_clientes.cell(row=ultimo,column=8).value = cidade
        sheet_clientes.cell(row=ultimo,column=9).value = telefone
        sheet_clientes.cell(row=ultimo,column=10).value = email
        sheet_clientes.cell(row=ultimo,column=11).value = aniversario
        sheet_clientes.cell(row=ultimo,column=12).value = cpf

    else:
        print('\nO cliente foi adicionado com sucesso!\n')
        return 1




__init__()
adicionarpedido()
